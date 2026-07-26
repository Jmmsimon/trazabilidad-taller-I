# -*- coding: utf-8 -*-
"""Genera Matriz_y_Casos_de_Prueba.xlsx limpio (matriz + 1 hoja por caso)."""
from pathlib import Path
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "entrega" / "Pruebas" / "Matriz_y_Casos_de_Prueba.xlsx"
DESKTOP = Path(r"c:\Users\JSIMON\Desktop") / OUT.name

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=10)
PASE_FILL = PatternFill("solid", fgColor="C6EFCE")
GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_matrix(wb):
    ws = wb.active
    ws.title = "Matrix de Pruebas"
    ws["B1"] = "Plan de pruebas — Plataforma de Trazabilidad Académica"
    ws["B1"].font = TITLE_FONT
    ws["B3"] = "Fecha de creación del plan de pruebas: 24/07/2026"
    ws["B3"].font = NORMAL
    ws["B4"] = "Plan de prueba creado por: Jean Marcos Meneses Simón"
    ws["B4"].font = NORMAL
    ws["B5"] = (
        "Proyecto: Trazabilidad Académica con Agentes de IA Generativa | "
        "Asesor: Ing. Walter Cueva"
    )
    ws["B5"].font = NORMAL

    headers = ["Caso de Prueba #", "Casos de Uso", "Area", "Estado", "Severidad", "Comentarios"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=7, column=i, value=h)
    style_header_row(ws, 7, 6)

    rows = [
        ("1", "Autenticación y control de acceso por roles", "Auth / RBAC", "—", 1, "Grupo funcional", True),
        ("1.1", "Inicio de sesión válido (estudiante)", "Auth / RBAC", "PASE", 1, "", False),
        ("1.2", "Inicio de sesión válido (docente)", "Auth / RBAC", "PASE", 1, "", False),
        ("1.3", "Inicio de sesión con credenciales inválidas", "Auth / RBAC", "PASE", 1, "", False),
        ("1.4", "Restricción de rutas por rol (estudiante no entra a admin)", "Auth / RBAC", "PASE", 1, "", False),
        ("1.5", "Cierre de sesión e invalidación de cookie", "Auth / RBAC", "PASE", 2, "", False),
        ("2", "Discovery multi-agente y aprobación de roadmap", "Discovery", "—", 1, "Grupo funcional", True),
        ("2.1", "Generación de propuesta técnica (drafter–validator–PO)", "Discovery", "PASE", 1, "", False),
        ("2.2", "Confirmación / iteración de propuesta por el estudiante", "Discovery", "PASE", 2, "", False),
        ("2.3", "Aprobación de roadmap por el docente", "Discovery", "PASE", 1, "", False),
        ("2.4", "Rechazo de roadmap con observaciones", "Discovery", "PASE", 2, "", False),
        ("2.5", "Persistencia de hitos y backlog_scrum en Firestore", "Discovery", "PASE", 1, "", False),
        ("3", "Tracking, AUTO-KANBAN y analítica", "Tracking", "—", 1, "Grupo funcional", True),
        ("3.1", "Configuración de repo_url y demo_url", "Tracking", "PASE", 1, "", False),
        ("3.2", "Ejecución de Analizar (Tracking deep)", "Tracking", "PASE", 1, "Piloto: completed", False),
        ("3.3", "AUTO-KANBAN: ítems con evidencia → done", "Tracking", "PASE", 1, "25/29 done", False),
        ("3.4", "Analítica estudiante (score, commits, competencias)", "Tracking", "PASE", 1, "Score 100; 15 commits; 80%", False),
        ("3.5", "Analítica docente (misma fuente de verdad)", "Tracking", "PASE", 1, "ADR-01", False),
        ("4", "Auditoría CSV vs código (solo reporte)", "Auditoría", "—", 1, "Grupo funcional", True),
        ("4.1", "Parseo CSV válido (coma / punto y coma)", "Auditoría", "PASE", 1, "También unitarias", False),
        ("4.2", "Rechazo de CSV vacío o sin columna título", "Auditoría", "PASE", 2, "", False),
        ("4.3", "Ejecución de auditoría y semáforo", "Auditoría", "PASE", 1, "", False),
        ("4.4", "Verificar que la auditoría NO altera Kanban alumno", "Auditoría", "PASE", 1, "ADR-02", False),
        ("4.5", "Consulta de status de auditoría (semáforo)", "Auditoría", "PASE", 2, "", False),
        ("5", "Administración, exportación y disponibilidad", "Admin / Ops", "—", 2, "Grupo funcional", True),
        ("5.1", "Listado y cambio de rol de usuarios", "Admin / Ops", "PASE", 2, "", False),
        ("5.2", "Exportación PDF/JSON del proyecto", "Admin / Ops", "PASE", 2, "", False),
        ("5.3", "Disponibilidad Swagger /docs del backend", "Admin / Ops", "PASE", 2, "", False),
        ("5.4", "Prueba de carga GET /docs (concurrencia)", "Admin / Ops", "PASE", 3, "Informe de carga LC-01/02/03", False),
        ("5.5", "Suite unitaria pytest (parser, schemas, semáforo)", "Admin / Ops", "PASE", 1, "41 passed", False),
    ]

    for i, (num, uso, area, estado, sev, com, is_group) in enumerate(rows):
        r = 8 + i
        for c, v in enumerate([num, uso, area, estado, sev, com], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name="Calibri", size=10, bold=is_group)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_group:
                cell.fill = GROUP_FILL
            elif estado == "PASE" and c == 4:
                cell.fill = PASE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

    autosize(ws, [16, 58, 14, 10, 12, 28])
    ws.row_dimensions[7].height = 22


CASOS = [
    {
        "nro": "1.1",
        "area": "Autenticación y control de acceso (RBAC)",
        "nombre": "Inicio de sesión válido — rol estudiante",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir la URL de la aplicación", "http://localhost:3000", "Se muestra el portal/formulario de acceso", "PASE"),
            ("Paso 2: Seleccionar o ingresar al flujo estudiante", "Portal estudiante", "Formulario de credenciales visible", "PASE"),
            ("Paso 3: Ingresar correo electrónico válido", "jmenesess1@upao.edu.pe", "Campo acepta el valor", "PASE"),
            ("Paso 4: Ingresar contraseña correcta", "Contraseña de prueba del entorno", "Campo enmascarado; listo para enviar", "PASE"),
            ("Paso 5: Confirmar inicio de sesión", "Click en botón de acceso", "Autenticación exitosa; cookie de sesión", "PASE"),
            ("Paso 6: Verificar redirección al dashboard", "URL resultante", "URL con /estudiante o equivalente", "PASE"),
            ("Paso 7: Verificar menú del rol", "Discovery / Kanban / Analítica", "Opciones de estudiante visibles", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Evidencia en Informe Capstone (resultados / anexos de pruebas).",
    },
    {
        "nro": "1.2",
        "area": "Autenticación y control de acceso (RBAC)",
        "nombre": "Inicio de sesión válido — rol docente",
        "rol": "Docente",
        "email": "waltercueva@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir la URL de la aplicación", "http://localhost:3000", "Portal de acceso visible", "PASE"),
            ("Paso 2: Ingresar credenciales de cuenta docente", "email + password docente", "Credenciales aceptadas por Firebase Auth", "PASE"),
            ("Paso 3: Confirmar login", "Click Ingresar", "Sesión creada (cookie httpOnly)", "PASE"),
            ("Paso 4: Verificar redirección al panel docente", "URL /docente", "Dashboard docente con proyectos", "PASE"),
            ("Paso 5: Verificar módulos docentes", "Hitos, Analítica, Auditoría CSV", "Módulos visibles según rol", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Evidencia en Informe Capstone y pruebas E2E (login docente).",
    },
    {
        "nro": "1.3",
        "area": "Autenticación y control de acceso (RBAC)",
        "nombre": "Inicio de sesión con credenciales inválidas",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir pantalla de login", "http://localhost:3000", "Formulario visible", "PASE"),
            ("Paso 2: Email válido y contraseña incorrecta", "password errónea", "Rechazo de acceso; mensaje de error", "PASE"),
            ("Paso 3: Intentar enviar con campos vacíos", "email/password vacíos", "Validación; no se autentica", "PASE"),
            ("Paso 4: Usuario inexistente", "noexiste@upao.edu.pe", "Error de autenticación; sin acceso", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Cubierto también en pruebas de caja negra (CN-02, CN-03).",
    },
    {
        "nro": "1.4",
        "area": "Autenticación y control de acceso (RBAC)",
        "nombre": "Restricción de rutas: estudiante no accede a administrador",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Iniciar sesión como estudiante", "Credenciales estudiante", "Dashboard estudiante", "PASE"),
            ("Paso 2: Navegar a ruta de administrador", "http://localhost:3000/administrador", "Acceso denegado o redirección", "PASE"),
            ("Paso 3: Verificar que no se exponen acciones admin", "UI admin", "Sin acciones de administración", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Severidad 1: falla de RBAC sería crítica.",
    },
    {
        "nro": "2.1",
        "area": "Discovery multi-agente",
        "nombre": "Generación de propuesta técnica (drafter–validator–PO)",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Acceder a inicio de proyecto", "Dashboard estudiante", "Formulario idea/stack disponible", "PASE"),
            ("Paso 2: Ingresar idea y stack", "Idea + Next.js/FastAPI", "Datos capturados en el request", "PASE"),
            ("Paso 3: Ejecutar Discovery", "Iniciar agentes", "Job en background; progreso en UI", "PASE"),
            ("Paso 4: Esperar finalización LangGraph", "drafter/validator/PO", "PropuestaTecnica generada", "PASE"),
            ("Paso 5: Revisar hitos y backlog", "UI roadmap", "Hitos y tareas visibles", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Evidencia en Informe Capstone (sección resultados / Discovery).",
    },
    {
        "nro": "2.3",
        "area": "Discovery multi-agente",
        "nombre": "Aprobación de roadmap por el docente",
        "rol": "Docente",
        "email": "waltercueva@upao.edu.pe",
        "pasos": [
            ("Paso 1: Iniciar sesión como docente", "Credenciales docente", "Panel docente", "PASE"),
            ("Paso 2: Abrir proyecto pendiente", "Proyecto piloto", "Detalle de propuesta visible", "PASE"),
            ("Paso 3: Ejecutar acción Aprobar", "Click Aprobar", "Estado actualizado a aprobado", "PASE"),
            ("Paso 4: Verificar persistencia", "Recargar detalle", "Aprobación se mantiene", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "",
    },
    {
        "nro": "3.1",
        "area": "Tracking / configuración de entregables",
        "nombre": "Configuración de repo_url y demo_url",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir configuración del proyecto", "Dashboard → configuración", "Campos repo y demo visibles", "PASE"),
            ("Paso 2: Ingresar URL de repositorio GitHub", "https://github.com/<org>/<repo>", "Valor aceptado", "PASE"),
            ("Paso 3: Ingresar URL de demo", "URL de despliegue", "Valor aceptado", "PASE"),
            ("Paso 4: Guardar configuración", "Click Guardar", "Persistencia en Firestore", "PASE"),
            ("Paso 5: Reabrir configuración", "—", "URLs guardadas se muestran", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Evidencia en Informe Capstone (configuración de entregables).",
    },
    {
        "nro": "3.2",
        "area": "Tracking deep",
        "nombre": "Ejecución de Analizar (pipeline Tracking)",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Verificar repo_url configurada", "Proyecto piloto", "URL presente", "PASE"),
            ("Paso 2: Pulsar botón Analizar", "UI Tracking", "Se inicia tracking; estado running", "PASE"),
            ("Paso 3: Esperar finalización", "GET tracking/status", "tracking_status = completed", "PASE"),
            ("Paso 4: Verificar score y commits", "UI / respuesta", "score_integridad y commits disponibles", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Piloto: completed. El tiempo depende de LLM + GitHub.",
    },
    {
        "nro": "3.3",
        "area": "AUTO-KANBAN",
        "nombre": "Actualización automática de estados Kanban con evidencia",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir tablero Kanban tras Analizar", "Vista backlog_scrum", "Columnas backlog/todo/in_progress/done", "PASE"),
            ("Paso 2: Identificar ítems con evidencia", "HU/EN con evidencia", "Ítems evidenciados en done", "PASE"),
            ("Paso 3: Contrastar conteo done vs total", "Piloto 29 ítems", "25/29 done en piloto", "PASE"),
            ("Paso 4: Verificar misma vista desde docente", "Panel docente Kanban", "Estados coherentes (ADR-01)", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Resultado piloto: 25/29 done (86.2%).",
    },
    {
        "nro": "3.4",
        "area": "Analítica",
        "nombre": "Analítica del estudiante — score, commits y competencias",
        "rol": "Estudiante",
        "email": "jmenesess1@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir módulo Analítica", "Dashboard estudiante", "Vista de métricas sin error", "PASE"),
            ("Paso 2: Verificar score de integridad", "Widget score", "Valor 0–100; piloto 100", "PASE"),
            ("Paso 3: Verificar commits analizados", "Lista/contador", "Piloto: 15 commits", "PASE"),
            ("Paso 4: Verificar % competencias", "Reporte competencias", "Piloto: 80%", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Evidencia en Informe Capstone (analítica piloto).",
    },
    {
        "nro": "4.3",
        "area": "Auditoría de backlog CSV",
        "nombre": "Ejecución de auditoría y obtención de semáforo",
        "rol": "Docente",
        "email": "waltercueva@upao.edu.pe",
        "pasos": [
            ("Paso 1: Abrir Auditoría de Avance", "Panel docente → proyecto", "UI de carga CSV visible", "PASE"),
            ("Paso 2: Seleccionar archivo CSV de backlog", "CSV export Notion / backlog", "Archivo aceptado", "PASE"),
            ("Paso 3: Ejecutar auditoría", "POST backlog-audit", "Job inicia; status consultable", "PASE"),
            ("Paso 4: Esperar resultado", "GET backlog-audit/status", "Semáforo y notas disponibles", "PASE"),
            ("Paso 5: Leer desviaciones del reporte", "UI semáforo", "Desviaciones listadas sin caída", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Evidencia en Informe Capstone (auditoría CSV).",
    },
    {
        "nro": "4.4",
        "area": "Auditoría de backlog CSV",
        "nombre": "Regla de negocio: auditoría solo-reporte (no altera Kanban)",
        "rol": "Docente / Estudiante",
        "email": "waltercueva@upao.edu.pe",
        "pasos": [
            ("Paso 1: Registrar estado Kanban del alumno", "Conteo por columna", "Baseline documentado", "PASE"),
            ("Paso 2: Ejecutar auditoría CSV hasta completar", "Como caso 4.3", "Semáforo disponible", "PASE"),
            ("Paso 3: Volver al Kanban del estudiante", "Rol estudiante", "Estados idénticos al baseline", "PASE"),
            ("Paso 4: Confirmar aviso de solo-reporte", "Banner/texto UI", "Mensaje coherente con ADR-02", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "solo_reporte=true; Kanban del alumno intacto.",
    },
    {
        "nro": "5.3",
        "area": "Disponibilidad del backend",
        "nombre": "Swagger UI /docs operativo",
        "rol": "Sistema / QA",
        "email": "—",
        "pasos": [
            ("Paso 1: Asegurar backend en ejecución", "uvicorn main:app --port 8000", "Proceso activo", "PASE"),
            ("Paso 2: Abrir http://127.0.0.1:8000/docs", "Navegador", "Swagger UI carga", "PASE"),
            ("Paso 3: Verificar presencia de endpoints", "Lista de rutas", "Rutas tracking y backlog-audit visibles", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "También verificado en pruebas de carga (LC-01/LC-02).",
    },
    {
        "nro": "5.5",
        "area": "Pruebas unitarias (caja blanca)",
        "nombre": "Ejecución suite pytest backend",
        "rol": "QA / Desarrollador",
        "email": "—",
        "pasos": [
            ("Paso 1: Activar entorno virtual del backend", "backend/.venv", "Entorno activo", "PASE"),
            ("Paso 2: Ejecutar pytest", "python -m pytest -v", "Recolección de tests en backend/tests", "PASE"),
            ("Paso 3: Verificar resultado final", "Resumen pytest", "41 passed; 0 failed", "PASE"),
        ],
        "problemas": "Ninguno",
        "comentarios": "Registro: Pruebas_unitarias/pytest_resultado.txt",
    },
]


def add_caso_sheet(wb, caso):
    w = wb.create_sheet(f"Caso de Prueba {caso['nro']}")
    w["A1"] = "Caso de prueba detallado — Plataforma de Trazabilidad Académica"
    w["A1"].font = TITLE_FONT
    meta = [
        ("Nro de Caso", caso["nro"]),
        ("Area de Prueba", caso["area"]),
        ("Nombre", caso["nombre"]),
        ("Rol Usado", caso["rol"]),
        ("Asignado A", "Jean Marcos Meneses Simón"),
        ("Login Email", caso["email"]),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        w.cell(row=i, column=1, value=k).font = LABEL_FONT
        w.cell(row=i, column=2, value=v).font = NORMAL
        w.cell(row=i, column=1).border = THIN
        w.cell(row=i, column=2).border = THIN

    header_row = 10
    cols = ["Pasos del caso de prueba", "Datos", "Resultado esperado", "Resultado de la prueba"]
    for c, h in enumerate(cols, 1):
        w.cell(row=header_row, column=c, value=h)
    style_header_row(w, header_row, 4)

    for i, (paso, datos, esp, res) in enumerate(caso["pasos"]):
        r = header_row + 1 + i
        for c, v in enumerate([paso, datos, esp, res], 1):
            cell = w.cell(row=r, column=c, value=v)
            cell.font = NORMAL
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 4 and v == "PASE":
                cell.fill = PASE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

    end = header_row + 1 + len(caso["pasos"])
    w.cell(row=end + 1, column=1, value="Problemas encontrados:").font = LABEL_FONT
    w.cell(row=end + 1, column=2, value=caso["problemas"]).font = NORMAL
    w.cell(row=end + 2, column=1, value="Otros comentarios:").font = LABEL_FONT
    w.cell(row=end + 2, column=2, value=caso["comentarios"]).font = NORMAL
    autosize(w, [48, 36, 42, 18])
    for r in range(header_row + 1, end):
        w.row_dimensions[r].height = 32


def main():
    wb = Workbook()
    build_matrix(wb)
    for caso in CASOS:
        add_caso_sheet(wb, caso)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    shutil.copy2(OUT, DESKTOP)

    wb2 = load_workbook(OUT)
    dirty = []
    for name in wb2.sheetnames:
        for row in wb2[name].iter_rows(values_only=True):
            for c in row:
                if c and any(x in str(c) for x in ("PEGAR", "████", "Adapte la Matriz", "Ejemplo de Matriz")):
                    dirty.append((name, str(c)[:80]))
    print("Hojas:", wb2.sheetnames)
    print("Casos detallados:", len(CASOS))
    print("Dirty:", dirty or "NINGUNO")
    print("OK", OUT)


if __name__ == "__main__":
    main()
